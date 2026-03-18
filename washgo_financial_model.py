"""
WashGo Financial Model Generator
Run:  python washgo_financial_model.py
Output: washgo_financial_model.xlsx
"""

import os
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import DataPoint

# ── Palette ────────────────────────────────────────────────────────────────────
GREEN_DARK   = "1B5E20"
GREEN_MED    = "2E7D32"
GREEN_LIGHT  = "E8F5E9"
GREEN_ALT    = "C8E6C9"
AMBER        = "FF6F00"
AMBER_LIGHT  = "FFF3E0"
NAVY         = "1A237E"
NAVY_LIGHT   = "E8EAF6"
RED_LIGHT    = "FFEBEE"
WHITE        = "FFFFFF"
GREY_HEADER  = "F3F4F6"
GREY_LIGHT   = "F9FAFB"
GREY_BORDER  = "E5E7EB"

OUTPUT_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                           "washgo_financial_model.xlsx")


# ── Style helpers ──────────────────────────────────────────────────────────────

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, color="000000", size=11, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic,
                name="Calibri")

def _border(style="thin"):
    side = Side(style=style, color=GREY_BORDER)
    return Border(left=side, right=side, top=side, bottom=side)

def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _header_style(ws, row, col, value, bg=GREEN_MED, fg=WHITE,
                   bold=True, size=11, h_align="center"):
    c = ws.cell(row=row, column=col, value=value)
    c.fill     = _fill(bg)
    c.font     = _font(bold=bold, color=fg, size=size)
    c.alignment = _align(h=h_align, v="center")
    c.border   = _border()
    return c

def _data_style(ws, row, col, value, bg=WHITE, bold=False, h_align="left",
                num_fmt=None, color="000000"):
    c = ws.cell(row=row, column=col, value=value)
    c.fill     = _fill(bg)
    c.font     = _font(bold=bold, color=color)
    c.alignment = _align(h=h_align, v="center")
    c.border   = _border()
    if num_fmt:
        c.number_format = num_fmt
    return c

def _section_header(ws, row, col, value, colspan=1):
    """Bold navy section header, no fill border, left-aligned."""
    c = ws.cell(row=row, column=col, value=value)
    c.font      = _font(bold=True, color=NAVY, size=12)
    c.fill      = _fill(NAVY_LIGHT)
    c.alignment = _align(h="left", v="center")
    c.border    = _border()
    return c

INR_FMT    = '₹#,##0'
INR_DEC    = '₹#,##0.00'
PCT_FMT    = '0.0%'
PCT_FMT2   = '0%'
NUM_FMT    = '#,##0'


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 1 — ASSUMPTIONS
# ══════════════════════════════════════════════════════════════════════════════

def build_assumptions(wb):
    ws = wb.create_sheet("Assumptions")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 48

    # Title row
    ws.row_dimensions[1].height = 36
    t = ws.cell(row=1, column=1,
                value="WashGo — Financial Model Assumptions")
    t.font      = _font(bold=True, color=WHITE, size=16)
    t.fill      = _fill(GREEN_MED)
    t.alignment = _align(h="center", v="center")
    ws.merge_cells("A1:C1")

    # Sub-header
    ws.row_dimensions[2].height = 20
    sub = ws.cell(row=2, column=1,
                  value="All values used as inputs for P&L Projection, Unit Economics, and Break-even sheets.")
    sub.font      = _font(italic=True, color="6B7280", size=10)
    sub.fill      = _fill(GREEN_LIGHT)
    sub.alignment = _align(h="center", v="center")
    ws.merge_cells("A2:C2")

    # Column headers
    for col, hdr in enumerate(["Parameter", "Value", "Notes / Formula"], 1):
        _header_style(ws, 3, col, hdr, bg=GREEN_DARK, size=11)

    assumptions = [
        # (Label, Value, Note, num_fmt)
        ("── ORDER VOLUME ──", None, None, None),
        ("Starting orders / month (Month 1)",        50,      "Seed-stage pilot assumption",              NUM_FMT),
        ("Monthly order growth rate — M1–6",         0.30,    "30% month-on-month growth in launch phase",PCT_FMT),
        ("Monthly order growth rate — M7–12",        0.20,    "20% as product-market fit improves",       PCT_FMT),
        ("Monthly order growth rate — M13–18",       0.15,    "15% steady-state scaling phase",           PCT_FMT),
        ("── REVENUE DRIVERS ──", None, None, None),
        ("Average order value (₹)",                  380,     "Blended across all 6 service types",       INR_FMT),
        ("WashGo commission %",                       0.25,    "25% of gross order value per transaction", PCT_FMT),
        ("Delivery fee per order (₹)",               49,      "Flat fee charged to customer per cycle",   INR_FMT),
        ("── SUBSCRIPTION REVENUE ──", None, None, None),
        ("Subscriber conversion rate",                0.03,    "3% of cumulative customers become subscribers", PCT_FMT),
        ("Avg subscription revenue / subscriber (₹)",1500,    "Blended across ₹999, ₹1,799, ₹2,999 plans", INR_FMT),
        ("── COST STRUCTURE ──", None, None, None),
        ("Fixed costs / month (₹)",                  120000,  "Salaries + rent + tech infrastructure",    INR_FMT),
        ("Variable cost per order (₹)",              85,      "Partner payout + packaging + misc",        INR_FMT),
        ("── CUSTOMER ACQUISITION ──", None, None, None),
        ("CAC — customer acquisition cost (₹)",      200,     "Blended digital + referral CAC",           INR_FMT),
        ("Monthly new customers (% of orders)",       0.10,    "10% of monthly orders = new unique users", PCT_FMT),
    ]

    row = 4
    for label, value, note, fmt in assumptions:
        ws.row_dimensions[row].height = 22
        if value is None:
            # Section sub-header
            c = ws.cell(row=row, column=1, value=label)
            c.font      = _font(bold=True, color=NAVY, size=10)
            c.fill      = _fill(NAVY_LIGHT)
            c.alignment = _align(h="left", v="center")
            c.border    = _border()
            for col in range(2, 4):
                d = ws.cell(row=row, column=col)
                d.fill   = _fill(NAVY_LIGHT)
                d.border = _border()
        else:
            bg = GREY_LIGHT if row % 2 == 0 else WHITE
            _data_style(ws, row, 1, label, bg=bg, h_align="left")
            v = _data_style(ws, row, 2, value, bg=bg, h_align="center",
                            num_fmt=fmt)
            _data_style(ws, row, 3, note, bg=bg, h_align="left",
                        color="4B5563")
        row += 1

    # Named ranges / freeze
    ws.freeze_panes = "A4"
    return ws


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 2 — P&L PROJECTION (18 months)
# ══════════════════════════════════════════════════════════════════════════════

def build_pl(wb):
    ws = wb.create_sheet("P&L Projection")
    ws.sheet_view.showGridLines = False

    col_widths = [10, 12, 18, 18, 20, 16, 16, 14, 14, 14, 18]
    col_letters = [get_column_letter(i+1) for i in range(len(col_widths))]
    for letter, width in zip(col_letters, col_widths):
        ws.column_dimensions[letter].width = width

    # Title
    ws.row_dimensions[1].height = 36
    t = ws.cell(row=1, column=1,
                value="WashGo — 18-Month P&L Projection")
    t.font      = _font(bold=True, color=WHITE, size=16)
    t.fill      = _fill(GREEN_MED)
    t.alignment = _align(h="center", v="center")
    ws.merge_cells("A1:K1")

    ws.row_dimensions[2].height = 18
    sub = ws.cell(row=2, column=1,
                  value="Figures in Indian Rupees (₹). Assumptions referenced from 'Assumptions' sheet.")
    sub.font      = _font(italic=True, color="6B7280", size=10)
    sub.fill      = _fill(GREEN_LIGHT)
    sub.alignment = _align(h="center", v="center")
    ws.merge_cells("A2:K2")

    # Column headers
    headers = [
        "Month", "Orders", "Commission\nRevenue (₹)",
        "Delivery\nRevenue (₹)", "Subscription\nRevenue (₹)",
        "Total\nRevenue (₹)", "Variable\nCosts (₹)",
        "Fixed\nCosts (₹)", "Total\nCosts (₹)",
        "EBITDA (₹)", "Cumulative\nCash (₹)"
    ]
    ws.row_dimensions[3].height = 36
    for col, hdr in enumerate(headers, 1):
        c = _header_style(ws, 3, col, hdr, bg=GREEN_MED, size=10)
        c.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=True)

    # Pull raw assumption values (hardcoded here, mirror the Assumptions sheet)
    start_orders  = 50
    growth_rates  = [0.30]*6 + [0.20]*6 + [0.15]*6   # M1–18 growth (applied to NEXT month)
    avg_order     = 380
    commission    = 0.25
    del_fee       = 49
    sub_conv      = 0.03
    avg_sub_rev   = 1500
    fixed_cost    = 120000
    var_cost      = 85

    orders = [start_orders]
    for i in range(17):
        orders.append(round(orders[-1] * (1 + growth_rates[i])))

    cumulative_cash = 0
    data_start_row  = 4

    for i, month_orders in enumerate(orders):
        row = data_start_row + i
        ws.row_dimensions[row].height = 20
        m_num = i + 1
        bg = GREY_LIGHT if m_num % 2 == 0 else WHITE

        # Subscriptions: cumulative customers * conversion
        cumulative_customers = sum(round(o * 0.10) for o in orders[:i+1])
        subscribers  = round(cumulative_customers * sub_conv)
        rev_comm     = round(month_orders * avg_order * commission)
        rev_del      = round(month_orders * del_fee)
        rev_sub      = round(subscribers * avg_sub_rev)
        total_rev    = rev_comm + rev_del + rev_sub
        var_costs    = round(month_orders * var_cost)
        total_costs  = var_costs + fixed_cost
        ebitda       = total_rev - total_costs
        cumulative_cash += ebitda

        row_data = [
            (f"Month {m_num}",  "left",  None,    "000000", False),
            (month_orders,      "right", NUM_FMT, "000000", False),
            (rev_comm,          "right", INR_FMT, "000000", False),
            (rev_del,           "right", INR_FMT, "000000", False),
            (rev_sub,           "right", INR_FMT, "000000", False),
            (total_rev,         "right", INR_FMT, GREEN_MED, True),
            (var_costs,         "right", INR_FMT, "000000", False),
            (fixed_cost,        "right", INR_FMT, "000000", False),
            (total_costs,       "right", INR_FMT, "EF5350", True),
            (ebitda,            "right", INR_FMT,
             GREEN_MED if ebitda >= 0 else "C62828", True),
            (cumulative_cash,   "right", INR_FMT,
             GREEN_MED if cumulative_cash >= 0 else "C62828", True),
        ]

        for col, (val, align, fmt, color, bold) in enumerate(row_data, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.fill      = _fill(bg)
            c.font      = _font(bold=bold, color=color)
            c.alignment = _align(h=align, v="center")
            c.border    = _border()
            if fmt:
                c.number_format = fmt

    # Totals row
    totals_row = data_start_row + 18
    ws.row_dimensions[totals_row].height = 24
    ws.cell(row=totals_row, column=1, value="TOTAL / AVERAGE").font = _font(bold=True, color=WHITE, size=11)
    ws.cell(row=totals_row, column=1).fill = _fill(GREEN_DARK)
    ws.cell(row=totals_row, column=1).alignment = _align(h="center", v="center")
    ws.cell(row=totals_row, column=1).border = _border()

    sum_cols = {
        2: (NUM_FMT, "000000"),   # orders
        3: (INR_FMT, WHITE),
        4: (INR_FMT, WHITE),
        5: (INR_FMT, WHITE),
        6: (INR_FMT, WHITE),
        7: (INR_FMT, WHITE),
        8: (INR_FMT, WHITE),
        9: (INR_FMT, WHITE),
        10: (INR_FMT, WHITE),
    }
    for col, (fmt, color) in sum_cols.items():
        col_letter = get_column_letter(col)
        c = ws.cell(row=totals_row, column=col,
                    value=f"=SUM({col_letter}{data_start_row}:{col_letter}{data_start_row+17})")
        c.fill           = _fill(GREEN_DARK)
        c.font           = _font(bold=True, color=WHITE)
        c.alignment      = _align(h="right", v="center")
        c.border         = _border()
        c.number_format  = fmt

    # Columns 10, 11 — leave empty in totals
    for col in [11]:
        c = ws.cell(row=totals_row, column=col, value="—")
        c.fill      = _fill(GREEN_DARK)
        c.font      = _font(bold=True, color=WHITE)
        c.alignment = _align(h="center", v="center")
        c.border    = _border()

    ws.freeze_panes = "B4"
    return ws


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 3 — UNIT ECONOMICS
# ══════════════════════════════════════════════════════════════════════════════

def build_unit_economics(wb):
    ws = wb.create_sheet("Unit Economics")
    ws.sheet_view.showGridLines = False

    col_widths = [22, 8, 8, 10, 16, 18, 16, 16, 16, 18, 14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Title
    ws.row_dimensions[1].height = 36
    t = ws.cell(row=1, column=1, value="WashGo — Unit Economics per Service")
    t.font      = _font(bold=True, color=WHITE, size=16)
    t.fill      = _fill(NAVY)
    t.alignment = _align(h="center", v="center")
    ws.merge_cells(f"A1:{get_column_letter(len(col_widths))}1")

    ws.row_dimensions[2].height = 18
    sub = ws.cell(row=2, column=1,
                  value="Per-order economics for each service type. Commission = 25% | Delivery fee = ₹49 | Partner payout = 65% of service price | Packaging = ₹15")
    sub.font      = _font(italic=True, color="6B7280", size=10)
    sub.fill      = _fill(NAVY_LIGHT)
    sub.alignment = _align(h="center", v="center")
    ws.merge_cells(f"A2:{get_column_letter(len(col_widths))}2")

    headers = [
        "Service", "Unit", "Avg Qty\nper Order",
        "Unit\nPrice (₹)", "Gross Order\nValue (₹)",
        "WashGo\nCommission (₹)", "Delivery\nFee (₹)",
        "Partner\nPayout (₹)", "Packaging\n(₹)",
        "WashGo Gross\nProfit / Order (₹)",
        "Gross\nMargin %",
    ]
    ws.row_dimensions[3].height = 36
    for col, hdr in enumerate(headers, 1):
        c = _header_style(ws, 3, col, hdr, bg=NAVY, size=10)
        c.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=True)

    services = [
        # (name,         unit,    avg_qty, unit_price)
        ("Regular Wash",  "kg",    5.0,    60),
        ("Express Wash",  "kg",    4.5,    100),
        ("Dry Cleaning",  "piece", 4,      150),
        ("Premium Care",  "piece", 3,      200),
        ("Ironing Only",  "piece", 10,     20),
        ("Wash & Iron",   "kg",    4.5,    80),
    ]

    commission_rate = 0.25
    delivery_fee    = 49
    partner_payout  = 0.65
    packaging       = 15

    for i, (name, unit, qty, price) in enumerate(services):
        row = 4 + i
        ws.row_dimensions[row].height = 22
        bg = GREEN_LIGHT if i % 2 == 0 else WHITE

        gross_order   = round(qty * price, 2)
        commission_amt= round(gross_order * commission_rate, 2)
        partner_pay   = round(gross_order * partner_payout, 2)
        gross_profit  = round(commission_amt + delivery_fee - packaging, 2)
        margin_pct    = gross_profit / (gross_order + delivery_fee) if (gross_order + delivery_fee) else 0

        row_vals = [
            (name,           "left",   None,    False, "000000"),
            (unit,           "center", None,    False, "6B7280"),
            (qty,            "center", "0.0",   False, "000000"),
            (price,          "center", INR_FMT, False, AMBER),
            (gross_order,    "center", INR_FMT, True,  AMBER),
            (commission_amt, "center", INR_FMT, True,  GREEN_MED),
            (delivery_fee,   "center", INR_FMT, False, "1565C0"),
            (partner_pay,    "center", INR_FMT, False, "EF5350"),
            (packaging,      "center", INR_FMT, False, "000000"),
            (gross_profit,   "center", INR_FMT, True,  GREEN_DARK),
            (margin_pct,     "center", PCT_FMT, True,  GREEN_MED),
        ]

        for col, (val, align, fmt, bold, color) in enumerate(row_vals, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.fill      = _fill(bg)
            c.font      = _font(bold=bold, color=color)
            c.alignment = _align(h=align, v="center")
            c.border    = _border()
            if fmt:
                c.number_format = fmt

    # Average row
    avg_row = 10
    ws.row_dimensions[avg_row].height = 24
    for col in range(1, len(col_widths)+1):
        c = ws.cell(row=avg_row, column=col)
        c.fill   = _fill(GREEN_LIGHT)
        c.font   = _font(bold=True, color=GREEN_DARK)
        c.border = _border()
    ws.cell(row=avg_row, column=1, value="AVERAGE / BLENDED").alignment = _align(h="center", v="center")

    for col_idx, fmt in [(5, INR_FMT), (6, INR_FMT), (10, INR_FMT), (11, PCT_FMT)]:
        col_letter = get_column_letter(col_idx)
        c = ws.cell(row=avg_row, column=col_idx,
                    value=f"=AVERAGE({col_letter}4:{col_letter}9)")
        c.fill           = _fill(GREEN_LIGHT)
        c.font           = _font(bold=True, color=GREEN_DARK)
        c.alignment      = _align(h="center", v="center")
        c.border         = _border()
        c.number_format  = fmt

    ws.freeze_panes = "A4"
    return ws


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 4 — BREAK-EVEN ANALYSIS
# ══════════════════════════════════════════════════════════════════════════════

def build_breakeven(wb):
    ws = wb.create_sheet("Break-even Analysis")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 48

    # Title
    ws.row_dimensions[1].height = 36
    t = ws.cell(row=1, column=1, value="WashGo — Break-even Analysis")
    t.font      = _font(bold=True, color=WHITE, size=16)
    t.fill      = _fill(AMBER)
    t.alignment = _align(h="center", v="center")
    ws.merge_cells("A1:C1")

    ws.row_dimensions[2].height = 18
    sub = ws.cell(row=2, column=1,
                  value="Monthly fixed cost breakdown and contribution margin calculation to determine break-even order volume.")
    sub.font      = _font(italic=True, color="6B7280", size=10)
    sub.fill      = _fill(AMBER_LIGHT)
    sub.alignment = _align(h="center", v="center")
    ws.merge_cells("A2:C2")

    # ── Section 1: Fixed Cost Breakdown ──
    row = 3
    _section_header(ws, row, 1, "SECTION 1 — Monthly Fixed Cost Breakdown")
    for col in range(2, 4):
        c = ws.cell(row=row, column=col)
        c.fill   = _fill(NAVY_LIGHT)
        c.border = _border()
        c.font   = _font(bold=True, color=NAVY)
    row += 1

    for col, hdr in enumerate(["Cost Item", "Monthly Amount (₹)", "Notes"], 1):
        _header_style(ws, row, col, hdr, bg=GREEN_MED, size=10)
    row += 1

    fixed_items = [
        ("Founder Salaries (2 co-founders)",  60000, "Early-stage modest draw"),
        ("Tech (hosting, tools, SaaS)",        15000, "AWS + Firebase + misc SaaS"),
        ("Operations Manager",                 25000, "1 FTE operations hire"),
        ("Marketing (digital + referral)",      15000, "Meta Ads + Google + referral"),
        ("Office / Miscellaneous",               5000, "Stationery, utilities, travel"),
    ]

    total_fixed = 0
    for label, amount, note in fixed_items:
        bg = GREY_LIGHT if row % 2 == 0 else WHITE
        _data_style(ws, row, 1, label, bg=bg)
        _data_style(ws, row, 2, amount, bg=bg, h_align="right",
                    num_fmt=INR_FMT, color=AMBER)
        _data_style(ws, row, 3, note, bg=bg, color="4B5563")
        total_fixed += amount
        row += 1

    # Total
    ws.row_dimensions[row].height = 24
    _data_style(ws, row, 1, "TOTAL MONTHLY FIXED COSTS", bg=GREEN_DARK,
                bold=True, color=WHITE)
    c = ws.cell(row=row, column=2, value=total_fixed)
    c.fill           = _fill(GREEN_DARK)
    c.font           = _font(bold=True, color=WHITE, size=12)
    c.alignment      = _align(h="right", v="center")
    c.border         = _border()
    c.number_format  = INR_FMT
    _data_style(ws, row, 3, "Sum of all recurring monthly costs",
                bg=GREEN_DARK, color=WHITE)
    row += 2

    # ── Section 2: Contribution Margin ──
    _section_header(ws, row, 1, "SECTION 2 — Contribution Margin per Order")
    for col in range(2, 4):
        c = ws.cell(row=row, column=col)
        c.fill   = _fill(NAVY_LIGHT)
        c.border = _border()
    row += 1

    for col, hdr in enumerate(["Metric", "Value (₹)", "Formula / Notes"], 1):
        _header_style(ws, row, col, hdr, bg=NAVY, size=10)
    row += 1

    avg_order   = 380
    commission  = 0.25
    del_fee     = 49
    var_cost    = 85

    rev_per_order  = round(avg_order * commission + del_fee, 2)
    contrib_margin = round(rev_per_order - var_cost, 2)
    breakeven_orders = round(total_fixed / contrib_margin)

    contrib_rows = [
        ("Average Gross Order Value",     avg_order,      "Blended across 6 service types"),
        ("× WashGo Commission (25%)",     round(avg_order * commission, 2),
                                          "= Avg order value × 25%"),
        ("+ Delivery Fee (flat)",         del_fee,        "Flat ₹49 per pickup + delivery"),
        ("= Total Revenue per Order",     rev_per_order,  "Commission + Delivery fee"),
        ("− Variable Cost per Order",     var_cost,       "Partner payout + packaging + ops"),
        ("= CONTRIBUTION MARGIN / ORDER", contrib_margin, "Revenue per order − Variable cost per order"),
    ]

    for label, val, note in contrib_rows:
        bg = GREEN_LIGHT if "CONTRIBUTION" in label else (GREY_LIGHT if row % 2 == 0 else WHITE)
        bold = "CONTRIBUTION" in label
        color = GREEN_DARK if "CONTRIBUTION" in label else "000000"
        _data_style(ws, row, 1, label, bg=bg, bold=bold, color=color)
        _data_style(ws, row, 2, val, bg=bg, bold=bold, h_align="right",
                    num_fmt=INR_FMT, color=GREEN_MED if bold else "000000")
        _data_style(ws, row, 3, note, bg=bg, color="4B5563")
        row += 1

    row += 1

    # ── Section 3: Break-even Calculation ──
    _section_header(ws, row, 1, "SECTION 3 — Break-even Order Volume")
    for col in range(2, 4):
        c = ws.cell(row=row, column=col)
        c.fill   = _fill(NAVY_LIGHT)
        c.border = _border()
    row += 1

    for col, hdr in enumerate(["Metric", "Value", "Explanation"], 1):
        _header_style(ws, row, col, hdr, bg=AMBER, fg=WHITE, size=10)
    row += 1

    breakeven_data = [
        ("Monthly Fixed Costs",
         total_fixed,
         "All costs that don't vary with order volume",
         INR_FMT, "000000", False),
        ("Contribution Margin per Order",
         contrib_margin,
         "Revenue earned per order minus direct variable costs",
         INR_FMT, GREEN_MED, True),
        ("Break-even Formula",
         "Fixed Costs ÷ Contribution Margin",
         f"₹{total_fixed:,} ÷ ₹{contrib_margin:,.0f} per order",
         None, NAVY, True),
        ("BREAK-EVEN ORDERS / MONTH",
         breakeven_orders,
         f"≈ Month 11 at current growth trajectory (1,600 orders/month)",
         NUM_FMT, WHITE, True),
    ]

    for label, val, note, fmt, color, bold in breakeven_data:
        bg = AMBER if "BREAK-EVEN ORDERS" in label else (GREY_LIGHT if row % 2 == 0 else WHITE)
        ws.row_dimensions[row].height = 26 if bold else 22
        _data_style(ws, row, 1, label, bg=bg, bold=bold,
                    color=WHITE if bg == AMBER else NAVY)
        c = ws.cell(row=row, column=2, value=val)
        c.fill      = _fill(bg)
        c.font      = _font(bold=bold, color=WHITE if bg == AMBER else color,
                            size=13 if "BREAK-EVEN" in label else 11)
        c.alignment = _align(h="right", v="center")
        c.border    = _border()
        if fmt:
            c.number_format = fmt
        _data_style(ws, row, 3, note, bg=bg,
                    color=WHITE if bg == AMBER else "4B5563", bold=bold)
        row += 1

    ws.freeze_panes = "A3"
    return ws


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 5 — FUNDING & RUNWAY
# ══════════════════════════════════════════════════════════════════════════════

def build_funding(wb):
    ws = wb.create_sheet("Funding & Runway")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 48

    # Title
    ws.row_dimensions[1].height = 36
    t = ws.cell(row=1, column=1, value="WashGo — Funding & Runway Analysis")
    t.font      = _font(bold=True, color=WHITE, size=16)
    t.fill      = _fill(NAVY)
    t.alignment = _align(h="center", v="center")
    ws.merge_cells("A1:D1")

    ws.row_dimensions[2].height = 18
    sub = ws.cell(row=2, column=1,
                  value="Seed round structure, burn rate, runway, and Series A milestone targets.")
    sub.font      = _font(italic=True, color="6B7280", size=10)
    sub.fill      = _fill(NAVY_LIGHT)
    sub.alignment = _align(h="center", v="center")
    ws.merge_cells("A2:D2")

    row = 3

    # ── Section 1: Seed Round Overview ──
    _section_header(ws, row, 1, "SECTION 1 — Seed Round Overview")
    for col in range(2, 5):
        c = ws.cell(row=row, column=col)
        c.fill = _fill(NAVY_LIGHT); c.border = _border()
    row += 1

    seed_overview = [
        ("Seed Round Amount",         "₹1,50,00,000", "₹1.5 Crore"),
        ("Instrument",                "SAFE / Convertible Note", "Standard SAFE with valuation cap"),
        ("Valuation Cap",             "₹8,00,00,000", "₹8 Crore post-money cap"),
        ("Discount Rate",             "20%", "Standard early-stage discount"),
        ("Use Period",                "18 Months", "Target: Series A readiness"),
        ("Target Series A Raise",     "₹8–10 Crore", "At ₹40 Crore valuation"),
    ]

    for col, hdr in enumerate(["Parameter", "Value", "Amount (₹)", "Notes"], 1):
        _header_style(ws, row, col, hdr, bg=GREEN_MED, size=10)
    row += 1

    for label, val, note in seed_overview:
        bg = GREY_LIGHT if row % 2 == 0 else WHITE
        _data_style(ws, row, 1, label, bg=bg)
        _data_style(ws, row, 2, val, bg=bg, h_align="center",
                    color=NAVY, bold=True)
        _data_style(ws, row, 3, "", bg=bg)
        _data_style(ws, row, 4, note, bg=bg, color="4B5563")
        row += 1

    row += 1

    # ── Section 2: Monthly Burn by Category ──
    _section_header(ws, row, 1, "SECTION 2 — Monthly Burn Rate by Category")
    for col in range(2, 5):
        c = ws.cell(row=row, column=col)
        c.fill = _fill(NAVY_LIGHT); c.border = _border()
    row += 1

    for col, hdr in enumerate(["Category", "Monthly Burn (₹)", "% of Total Burn",
                                "Notes"], 1):
        _header_style(ws, row, col, hdr, bg=AMBER, fg=WHITE, size=10)
    row += 1

    burn_categories = [
        ("Tech & Product",              31250, "25% of ₹1.5Cr ÷ 40 months (15% blended)"),
        ("Operations & Logistics",      37500, "30% allocation · partner & logistics costs"),
        ("Marketing & Acquisition",     31250, "25% allocation · CAC + brand building"),
        ("Team Salaries (3 hires)",     18750, "15% allocation · ops mgr + 2 field staff"),
        ("Working Capital & Misc",       6250, "5% allocation · deposits, buffer, misc"),
    ]

    total_burn = sum(b for _, b, _ in burn_categories)

    for label, burn, note in burn_categories:
        bg = GREY_LIGHT if row % 2 == 0 else WHITE
        pct = burn / total_burn
        _data_style(ws, row, 1, label, bg=bg)
        _data_style(ws, row, 2, burn, bg=bg, h_align="right",
                    num_fmt=INR_FMT, color=AMBER, bold=True)
        _data_style(ws, row, 3, pct, bg=bg, h_align="center",
                    num_fmt=PCT_FMT, color=NAVY)
        _data_style(ws, row, 4, note, bg=bg, color="4B5563")
        row += 1

    # Total burn row
    ws.row_dimensions[row].height = 24
    _data_style(ws, row, 1, "TOTAL MONTHLY BURN", bg=GREEN_DARK,
                bold=True, color=WHITE)
    c = ws.cell(row=row, column=2, value=total_burn)
    c.fill = _fill(GREEN_DARK); c.font = _font(bold=True, color=WHITE, size=12)
    c.alignment = _align(h="right", v="center"); c.border = _border()
    c.number_format = INR_FMT
    _data_style(ws, row, 3, 1.0, bg=GREEN_DARK, color=WHITE,
                num_fmt=PCT_FMT, bold=True)
    _data_style(ws, row, 4, "Blended average monthly cash out across 18 months",
                bg=GREEN_DARK, color=WHITE, bold=True)
    row += 2

    # ── Section 3: Runway ──
    seed_amount = 15000000
    runway_months = round(seed_amount / total_burn)

    _section_header(ws, row, 1, "SECTION 3 — Runway Calculation")
    for col in range(2, 5):
        c = ws.cell(row=row, column=col)
        c.fill = _fill(NAVY_LIGHT); c.border = _border()
    row += 1

    runway_data = [
        ("Seed Amount",             seed_amount, INR_FMT, "₹1.5 Crore raised"),
        ("Average Monthly Burn",    total_burn,  INR_FMT, "Blended burn across all categories"),
        ("Runway (months)",         runway_months, NUM_FMT, f"= ₹{seed_amount/100000:.0f}L ÷ ₹{total_burn/1000:.0f}K/month"),
    ]

    for col, hdr in enumerate(["Metric", "Value", "", "Notes"], 1):
        _header_style(ws, row, col, hdr, bg=GREEN_MED, size=10)
    row += 1

    for label, val, fmt, note in runway_data:
        bg = GREY_LIGHT if row % 2 == 0 else WHITE
        bold = "Runway" in label
        color = GREEN_DARK if bold else "000000"
        _data_style(ws, row, 1, label, bg=bg, bold=bold, color=color)
        c = ws.cell(row=row, column=2, value=val)
        c.fill = _fill(bg); c.font = _font(bold=bold, color=GREEN_MED if bold else AMBER)
        c.alignment = _align(h="right", v="center"); c.border = _border()
        c.number_format = fmt
        _data_style(ws, row, 3, "", bg=bg)
        _data_style(ws, row, 4, note, bg=bg, color="4B5563")
        row += 1

    row += 1

    # ── Section 4: Series A Milestone Targets ──
    _section_header(ws, row, 1, "SECTION 4 — Series A Milestone Targets")
    for col in range(2, 5):
        c = ws.cell(row=row, column=col)
        c.fill = _fill(NAVY_LIGHT); c.border = _border()
    row += 1

    for col, hdr in enumerate(["Milestone", "Target", "By Month", "Notes"], 1):
        _header_style(ws, row, col, hdr, bg=NAVY, fg=WHITE, size=10)
    row += 1

    milestones = [
        ("Orders per Month",    "5,000+",    "Month 18",  "Full Hyderabad coverage"),
        ("Monthly Revenue",     "₹25L+",     "Month 18",  "₹25 Lakh MRR target"),
        ("EBITDA Positive",     "3+ months", "Month 16+", "Sustained operational profitability"),
        ("Active Subscribers",  "900+",      "Month 18",  "Predictable recurring revenue base"),
        ("City Coverage",       "All Hyderabad", "Month 18", "All major IT corridors + residential"),
        ("Partner Facilities",  "15+",       "Month 15",  "Capacity for 5K+ orders/month"),
        ("Series A Raise",      "₹8–10 Cr",  "Month 18+", "At ₹40 Crore post-money valuation"),
    ]

    for label, target, timeline, note in milestones:
        bg = GREY_LIGHT if row % 2 == 0 else WHITE
        _data_style(ws, row, 1, label, bg=bg, bold=True, color=NAVY)
        _data_style(ws, row, 2, target, bg=bg, h_align="center",
                    color=GREEN_DARK, bold=True)
        _data_style(ws, row, 3, timeline, bg=bg, h_align="center",
                    color=AMBER, bold=True)
        _data_style(ws, row, 4, note, bg=bg, color="4B5563")
        row += 1

    ws.freeze_panes = "A3"
    return ws


# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════

def main():
    wb = Workbook()
    # Remove default blank sheet
    wb.remove(wb.active)

    print("Building Sheet 1: Assumptions ...")
    build_assumptions(wb)

    print("Building Sheet 2: P&L Projection ...")
    build_pl(wb)

    print("Building Sheet 3: Unit Economics ...")
    build_unit_economics(wb)

    print("Building Sheet 4: Break-even Analysis ...")
    build_breakeven(wb)

    print("Building Sheet 5: Funding & Runway ...")
    build_funding(wb)

    wb.save(OUTPUT_PATH)
    print(f"\nFinancial model saved: washgo_financial_model.xlsx")
    print(f"Full path: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
